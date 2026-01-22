"""python3 SMS automation script using pyautogui and Microsoft Link to Windows
This script automates SMS sending through PC/Laptop using Google Messages Web and Python scripting
The script first takes excel file path as arguments containing columns named
"Mobile Number" and "Message". Mobile number columns containing all the mobile numbers
and Message containing customised message corresponding to each mobile number.
It's important to write columns names with correct spelling.
First setup your sms sending account in messages.google.com
The user then needs to run this program in terminal/powershell/DOS as following
python3 autoSMSMessage.py path_toexcel.xlsx
After prompt message saying switch to  messages.google.com
in 2 seconds.
The line number
The script requires google messages app enabled phone and linking it on web

"""

import sys
import time
import pandas as pd
import pyautogui
import pyperclip
import os
from openpyxl.utils import get_column_letter
import openpyxl

# Load data - converting all to string and filling empty cells avoids TypeErrors
global file_path
file_path = sys.argv[
    1
]  # Provide your Excel file path in command like python3 % file.xlsx
df = pd.read_excel(file_path).fillna("").astype(str)
output_file_name = os.path.basename(file_path)
df.to_excel(f"backup{output_file_name}", index=False)


def autofit_existing_excel(filename):
    # Load the existing workbook
    wb = openpyxl.load_workbook(filename)

    for sheetname in wb.sheetnames:
        ws = wb[sheetname]

        # Iterate through each column in the sheet
        for col in ws.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)  # Get 'A', 'B', etc.

            for cell in col:
                try:
                    # Check length of cell value as a string
                    if cell.value:
                        if (
                            len(str(cell.value)) > max_length
                            and len(str(cell.value)) < 40
                        ):
                            max_length = len(str(cell.value))
                except FileNotFoundError:
                    pass

            # Apply width with a small margin (e.g., 1.2x length + 2)
            adjusted_width = (max_length + 1) * 1.1
            final_width = min(adjusted_width, 40)
            ws.column_dimensions[column_letter].width = final_width

    # Save the changes (can overwrite or save as new)
    wb.save(filename)
    print(f"Autofit complete for all sheets in {filename}")


def run_automation(df):
    pyautogui.FAILSAFE = True
    print("Switch to your target page. Autostarting in 2 seconds...")
    time.sleep(2)
    count = 0
    try:
        for index, row in df.iterrows():
            # Correctly accessing single cell values from the row
            count += 1
            if df.at[index, "smsStatus"] == "SMS sent":
                print(index, " already sent")
                count -= 1
                continue
            elif "smsStatus" not in df.columns:
                df["smsStatus"] = None
            elif count <= 100:
                mobile = row["Mobile_Number"]
                message = row["Message"]
                # Step 1: Copy/Paste Mobile Number
                pyautogui.click(x=187, y=150)
                time.sleep(0.3)
                pyautogui.hotkey("ctrl", "k")
                # starts a new chat
                time.sleep(0.5)
                pyperclip.copy(str(mobile))
                time.sleep(0.5)
                pyautogui.hotkey("ctrl", "v")
                time.sleep(0.5)
                # Step 2: Navigate to Message field
                pyautogui.press("enter")
                time.sleep(0.9)
                pyautogui.press("enter")
                time.sleep(1.5)

                # Step 3: Copy/Paste Message
                pyperclip.copy(str(message))
                time.sleep(0.3)
                pyautogui.hotkey("ctrl", "v")
                time.sleep(0.3)
                pyautogui.press("enter")
                time.sleep(1.9)
                print(f"{count}th SMS sent successfully to Mobile {mobile} ")
                # df = df.drop(index)
                df.at[index, "smsStatus"] = "SMS sent"
            else:
                print("100 messages sent,exiting")
                df.to_excel(output_file_name, index=False)
                autofit_existing_excel(output_file_name)
                break

        print("Sending SMS completed")
        df.to_excel(output_file_name, index=False)
        autofit_existing_excel(output_file_name)

    except pyautogui.FailSafeException:
        print("Exiting")
        df.to_excel(output_file_name, index=False)
        autofit_existing_excel(output_file_name)
        sys.exit()
    except KeyboardInterrupt:
        print("Exiting")
        df.to_excel(output_file_name, index=False)
        autofit_existing_excel(output_file_name)
        sys.exit()


if __name__ == "__main__":
    run_automation(df)
