#!/usr/bin/env python3
"""
copyNumbersAndAppendExcel.py
- Reads text from the system clipboard
- Finds all distinct 8-digit numbers
- Appends them (one per row) to Sheet2 of the specified Excel .xlsx file
Usage:
    python copyNumbersAndAppendExcel.py /path/to/file.xlsx
"""

import re
import sys
from pathlib import Path
import time
import pyperclip
import pyautogui
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def find_8_digit_numbers(text):
    # match standalone or within other characters but ensure exactly 8 consecutive digits, change 8 to any number text you want
    return re.findall(r"(?<!\d)(\d{8})(?!\d)", text)


def find_name_of_sheet(text):
    find = re.search(r"Transaction Details for FPS \d{7}\b", text)
    if find:
        message = find.group()
        fps_code = re.findall(r"\d{7}", message)
        return fps_code
    else:
        return "no match found"

from openpyxl import load_workbook
from pathlib import Path

def append_to_sheet2(xlsx_path, values, fps_code):
    xlsx = Path(xlsx_path)
    if not xlsx.exists():
        raise FileNotFoundError(f"File not found: {xlsx_path}")
    
    # Use read_only=False (default) but keep data_only=True if you don't need formulas
    wb = load_workbook(filename=str(xlsx))
    
    if "Sheet2" in wb.sheetnames:
        ws = wb["Sheet2"]
    else:
        ws = wb.create_sheet("Sheet2")

    # 1. Faster 'is_empty' check
    # Instead of tuple(iter_rows), just check if A1 has a value
    is_empty = ws.max_row == 1 and ws['A1'].value is None

    # 2. Prepare data in bulk
    # Converting values here is faster than doing it inside the Excel write loop
    fps_val = int(fps_code[0])
    
    # 3. Use .append() for speed
    # This is much faster than ws.cell(row=i, column=j)
    for val in values:
        ws.append([fps_val, int(val)])

    wb.save(str(xlsx))
    wb.close() # Always close to free up memory
    
def append_to_sheet(xlsx_path, values, fps_code):
    xlsx = Path(xlsx_path)
    if not xlsx.exists():
        raise FileNotFoundError(f"File not found: {xlsx_path}")
    wb = load_workbook(filename=str(xlsx))
    # ensure Sheet2 exists (create if not)
    if "Sheet2" in wb.sheetnames:
        ws = wb["Sheet2"]
    else:
        ws = wb.create_sheet("Sheet2")
# 1. Prepare your data in memory as a list of tuples
    start_row = ws.max_row + 1 if any(tuple(ws.iter_rows(min_row=1, max_row=1))) else 1
    for i, val in enumerate(values, start=start_row):
        ws.cell(row=i, column=2, value=int(val))
        ws.cell(row=i, column=1, value=int(fps_code[0]))
    wb.save(str(xlsx))


def main():
   # if len(sys.argv) != 2:
   #     print("Usage: python append_8digit_to_sheet2.py /path/to/file.xlsx")
   #     sys.exit(1)
   # xlsx_path = sys.argv[1]
    xlsx_path = r"Shesh parivar suchi January 2026.xlsx"
    text = pyperclip.paste()
    if not text:
        print("Clipboard is empty.")
        sys.exit(1)
    found = find_8_digit_numbers(text)
    if not found:
        print("No 8-digit numbers found in clipboard text.")
        sys.exit(0)
    # keep order and remove duplicates while preserving order
    start_time = time.perf_counter()
    seen = set()
    fps_code = find_name_of_sheet(text)
    print(fps_code)
    pyautogui.click(x=574, y=557)
    time.sleep(.5)
    pyautogui.hotkey("home",presses=2)
    time.sleep(.5)
    pyautogui.press('pagedown', presses=3)
   
    pyautogui.press('pagedown',int(abs(int(fps_code[0]))%100/20-1))
    
    unique = [x for x in found if not (x in seen or seen.add(x))]
    append_to_sheet2(xlsx_path, unique, fps_code)
    print(f"Appended {len(unique)} value(s) to Sheet2 of {xlsx_path}.")
    end = time.perf_counter()
    duration = -start_time + end
    print(f"Time taken  : {duration:.1f} seconds")
    


if __name__ == "__main__":
    main()
