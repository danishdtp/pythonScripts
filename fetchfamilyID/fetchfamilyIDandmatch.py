"""
Opens specified excel in argument
Copies 9 digit Id
Opens a page of web_url
Clicks on a text box
Paste the 9 digit Id
Clicks on submit button
Copies 10 digit ID
"""

import pandas as pd
import pyperclip
import pyautogui
import time
import re
import sys
import os
import openpyxl
from openpyxl.utils import get_column_letter
# Step 1: Load the Excel file into a DataFrame


def load_excel(file_path):
    return pd.read_excel(file_path)


def autofit_existing_excel(filename):
    # Load the existing workbook
    wb = openpyxl.load_workbook(filename)

    for sheetname in wb.sheetnames:
        ws = wb[sheetname]

        # Iterate through each column in the sheet
        for col in ws.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)  # Get 'A', 'B', etc.

            for cell in col:
                try:
                    # Check length of cell value as a string
                    if cell.value:
                        if (
                            len(str(cell.value)) > max_length
                            and len(str(cell.value)) < 20
                        ):
                            max_length = len(str(cell.value))
                except:
                    pass

            # Apply width with a small margin (e.g., 1.2x length + 2)
            adjusted_width = (max_length + 1) * 1.1
            final_width = min(adjusted_width, 20)
            ws.column_dimensions[column_letter].width = final_width

    # Save the changes (can overwrite or save as new)
    wb.save(filename)
    print(f"Autofit complete for all sheets in {filename}")


# Step 2: Identify all 8-digit numbers in the DataFrame
def find_eight_digit_cells(df):
    eight_digit_cells = []
    for index, row in df.iterrows():
        for col in df.columns:
            value = str(row[col])  # Convert to string for checking
            value = value.split(".")[0]
            if len(value) == 9 and value.isdigit():
                eight_digit_cells.append((index, col, value))
    return eight_digit_cells


# Step 3: Process each 8-digit number and perform required actions
def process_eight_digit_numbers(df, cells, output_file_name):
    total_count = len(df)
    print("Total : ", total_count)
    count = 0
    start_time = time.perf_counter()
    print("Script starting in 2 seconds switch to browser")
    time.sleep(1)
    global estimate

    def clicking_Function(value):
        time.sleep(0.2)
        pyautogui.click(x=952, y=534)  # Replace with actual coordinates
        time.sleep(0.1)
        pyautogui.click(x=952, y=534)  # Replace with actual coordinates
        time.sleep(0.2)
        pyautogui.typewrite(value)
        pyautogui.hotkey("tab")
        time.sleep(1)
        pyautogui.hotkey("tab")
        time.sleep(0.2)
        pyautogui.hotkey("enter")
        text = "Enter characters being displayed in above image"
        for i in range(30):
            time.sleep(0.5)
            # 3) Paste value in a search field using pyautogui
            pyautogui.hotkey("ctrl", "a")
            time.sleep(0.1)
            pyautogui.hotkey("ctrl", "c")
            time.sleep(0.1)
            text = pyperclip.paste()
            if (
                "Please correct/select/enter proper data highlighting" in text
                or "Warning" in text
                or "Invalid Data" in text
                or "Wrong Captcha Code, Please enter Correct Captcha Code" in text
                or "Enter Valid Member ID!" in text
                or "Enter characters being displayed in above image." not in text
            ):
                break
        return text

    try:
        for index, col, value in cells:
            # 1) Copy the 8-digit number to clipboard
            count += 1
            print(f"S.No {count}/{total_count}", end=" ")
            pattern = r"\b\d{8}\b"
            if "updatedSamagra" in df.columns:
                dataCheck = df.at[index, "updatedSamagra"]
                if re.search(pattern, str(dataCheck)) or "id not found" in str(
                    dataCheck
                ):
                    print("Value already exists")
                    if count == 1:
                        estimate = 1
                    continue
            pyperclip.copy(value)
            check = "start"
            for i in range(60):
                if "Enter characters being displayed in above image" not in check:
                    if i == 30 or i == 0:
                        time.sleep(0.2)  # Wait for a second
                        web_url = "https://csmsmpscsc.mp.gov.in/rationmitra/EBS/RCMS/AddMember.aspx"  # Change to the desired URL
                        pyautogui.hotkey("ctrl", "l")  # Focus address bar
                        time.sleep(0.2)  # Wait for a second
                        pyautogui.write(web_url)
                        pyautogui.press("enter")
                        time.sleep(0.2)
                    time.sleep(0.3)
                    pyautogui.hotkey("ctrl", "a")
                    time.sleep(0.1)
                    pyautogui.hotkey("ctrl", "c")
                    check = pyperclip.paste()
                else:
                    check = "start"
                    break
            copied_text = clicking_Function(value)

            while (
                "Please correct/select/enter proper data highlighting" in copied_text
                or "Wrong Captcha Code, Please enter Correct Captcha Code"
                in copied_text
                or "Enter Valid Member ID!" in copied_text
            ):
                pyautogui.click(x=963, y=728)
                copied_text = clicking_Function(value)
            match = re.search(r"\b\d{8}\b", copied_text)
            if match:
                found_number = match.group(0)
                # Add the found digit number into a new column in the DataFrame
                df.loc[index, "updatedSamagra"] = int(found_number)
                print(f"{count}/{total_count} - {value},  - {found_number}")

            elif "Member details not found for above" in copied_text:
                df.loc[index, "updatedSamagra"] = "ID not found"
                print(f"RC Number {value} is not valid")

            else:
                df.loc[index, "updatedSamagra"] = None
                print(f"{count}/{total_count} - {value},  - NA")

        return df
    except pyautogui.PyAutoGUIException or KeyboardInterrupt:
        print("Failsafe triggered, writing results to excel files")
        df.to_excel(output_file_name, index=False)
        autofit_existing_excel(output_file_name)
        sys.exit(0)


# Main function to execute the script
def main():
    file_path = sys.argv[1]  # Provide your Excel file path
    df = load_excel(file_path)
    output_file_name = os.path.basename(file_path)
    # backup_file
    df.to_excel(f"backup{output_file_name}", index=False)
    print("backup file created")
    print(output_file_name)
    eight_digit_cells = find_eight_digit_cells(df)
    if eight_digit_cells:
        df = process_eight_digit_numbers(df, eight_digit_cells, output_file_name)
        # df.to_excel(output_file_name, index=False)  # Output file name
        if df is not None and not df.empty:
            df.to_excel(output_file_name, index=False)
            autofit_existing_excel(output_file_name)
        else:
            print("DataFrame is empty or invalid, skipping export.")
        print(f"Saved to {output_file_name}")
    else:
        print("no 8 digit found")


if __name__ == "__main__":
    main()
