import pandas as pd
import os
import sys
import openpyxl
from openpyxl.utils import get_column_letter

# Configuration
input_file = sys.argv[1]
column_to_split_on = "LB"
# Replace with your actual column name
output_dir = "OutputFiles"


def autofit_existing_excel(filename):
    # Load the existing workbook
    wb = openpyxl.load_workbook(filename)

    for sheetname in wb.sheetnames:
        ws = wb[sheetname]

        # Iterate through each column in the sheet
        for col in ws.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)  # Get 'A', 'B', etc.

            for cell in col:
                try:
                    # Check length of cell value as a string
                    if cell.value:
                        if (
                            len(str(cell.value)) > max_length
                            and len(str(cell.value)) < 20
                        ):
                            max_length = len(str(cell.value))
                except:
                    pass

            # Apply width with a small margin (e.g., 1.2x length + 2)
            adjusted_width = (max_length + 1) * 1.1
            final_width = min(adjusted_width, 20)
            ws.column_dimensions[column_letter].width = final_width

    # Save the changes (can overwrite or save as new)
    wb.save(filename)
    print(f"Autofit complete for all sheets in {filename}")


# Create output directory if it doesn't exist
if not os.path.exists(output_dir):
    os.makedirs(output_dir)

# Read the excel file
df = pd.read_excel(input_file)
df.columns = df.columns.str.strip()  # Clean up column names

# Group by the specified column and save to separate files
for unique_value, group_data in df.groupby(column_to_split_on):
    # Sanitize the filename in case of problematic characters
    output_filename = f"{output_dir}/{unique_value}.xlsx".replace("/", "_").replace(
        "\\", "_"
    )
    print(output_filename)
    fileNamePath = f"{output_dir}/{output_filename}"
    group_data.to_excel(fileNamePath, index=False)
    autofit_existing_excel(fileNamePath)
    print(f"Created file: {output_filename}")
