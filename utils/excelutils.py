from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ExcelFormatter:
    """A module for standardizing Excel file formatting using OpenPyXL."""

    def __init__(self, file_path):
        self.file_path = file_path
        self.wb = load_workbook(file_path)
        self.ws = self.wb.active
        # Define a standard thin border style
        self.thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )

    def apply_standard_format(self, max_width=30, header_range='1:1'):
        """Applies all standard formatting: autofit, wrap, page fit, headers, and borders."""
        self._apply_cell_styles(max_width)
        self._fit_to_one_page()
        self._repeat_headers(header_range)

    def _apply_cell_styles(self, max_width):
        """Autofits columns, enables text wrapping, and adds borders."""
        for col_cells in self.ws.columns:
            # FIX: col_cells is a tuple of cells. Get the column index from the first cell.
            first_cell = col_cells[0]
            column_letter = get_column_letter(first_cell.column)
            
            max_length = 0
            for cell in col_cells:
                # Apply styles to every cell
                cell.alignment = Alignment(wrap_text=True)
                cell.border = self.thin_border
                
                # Calculate max content length
                if cell.value:
                    try:
                        length = len(str(cell.value))
                        if length > max_length:
                            max_length = length
                    except:
                        pass
            
            # Set the column width based on calculation
            adjusted_width = min(max_length + 2, max_width)
            self.ws.column_dimensions[column_letter].width = adjusted_width


    def _fit_to_one_page(self):
        """3. Fit all columns on one page."""
        self.ws.sheet_properties.pageSetUpPr.fitToPage = True
        self.ws.page_setup.fitToWidth = 1
        self.ws.page_setup.fitToHeight = 0 

    def _repeat_headers(self, header_range):
        """4. Repeat header rows for printing."""
        self.ws.print_title_rows = header_range

    def save(self, output_path=None):
        path = output_path or self.file_path
        self.wb.save(path)
        return path

def format_file(input_file, output_file=None):
    """Utility function to format a file in one line."""
    formatter = ExcelFormatter(input_file)
    formatter.apply_standard_format()
    return formatter.save(output_file)
